import json
import os
import queue
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, wait, ALL_COMPLETED

import requests as requests
from loguru import logger
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Side, Border
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.exceptions import InsecureRequestWarning

TIME_FORM = '%Y_%m_%d_%H_%M_%S'

# 配置文件所在目录
CONFIGDIR = 'config'

INPUTDIR = 'inputFile'
OUTPUTDIR = 'outputFile'
LOGDIR = 'log'

REQUEST_HEADERS = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/100.0.4896.60 Safari/537.36 Edg/100.0.1185.29 ',
    # 'Host': 'zhannei.baidu.com',
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,'
              'application/signed-exchange;v=b3;q=0.9',
    'accept-encoding': 'gzip, deflate, br',
    # 'content-type': 'image/jpeg'
}

PROXY = 'localhost:14978'
PROXIES = {
    'http': 'http://' + PROXY,
    'https': 'http://' + PROXY
}

THREAD_COUNT = 2
# 忽略requests证书警告
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

tpe = None
task_list = []


class BoundThreadPollExecutor(ThreadPoolExecutor):
    def __init__(self, *args, **kwargs):
        super(BoundThreadPollExecutor, self).__init__(*args, **kwargs)
        self._work_queue = queue.Queue(kwargs['max_workers'])


# task_list = []

def set_configuration():
    global TIME_FORM
    global INPUTDIR
    global OUTPUTDIR
    global LOGDIR
    global REQUEST_HEADERS
    global PROXIES
    global THREAD_COUNT
    global tpe
    config_filepath = CONFIGDIR + '/' + 'configuration.json'
    config_jsonfile = open(config_filepath, 'r', encoding='utf-8')
    config_json = json.load(config_jsonfile)
    TIME_FORM = config_json['TIME_FORM']
    INPUTDIR = config_json['INPUTDIR']
    OUTPUTDIR = config_json['OUTPUTDIR']
    LOGDIR = config_json['LOGDIR']
    REQUEST_HEADERS = config_json['REQUEST_HEADERS']
    PROXIES = config_json['PROXIES']
    THREAD_COUNT = config_json['THREAD_COUNT']
    tpe = BoundThreadPollExecutor(max_workers=THREAD_COUNT)
    # tpe = ThreadPoolExecutor(max_workers=THREAD_COUNT)


def make_filedir():
    dirs = [INPUTDIR, OUTPUTDIR, LOGDIR]
    for dirpath in dirs:
        if not os.path.exists(dirpath):
            os.mkdir(dirpath)


def get_filepath(filename_full):
    """
    获取输入输出文件相对路径
    :param filename_full: 文件全名
    :return: （输入文件路径，输出文件路径）
    """
    # 获取前缀（文件名称）
    filename_tup = os.path.splitext(filename_full)
    filename = filename_tup[0]
    filename_suffix = filename_tup[1]
    time.localtime(time.time())
    time_str = time.strftime(TIME_FORM, time.localtime(time.time()))
    return INPUTDIR + '/' + filename_full, OUTPUTDIR + '/' + filename + '_' + time_str + filename_suffix, LOGDIR + '/' + filename + '_' + time_str + '.log'


def get_highlight_style() -> NamedStyle:
    """
    生成标注对象
    :return:
    """
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True, color="ff0100")
    # 背景填充
    highlight.fill = PatternFill("solid", fgColor="DDDDDD")
    bd = Side(style='thick', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    return highlight


def cell_checker(value) -> bool | int:
    """
    审查内容
    :param value:
    :return:
    """
    value_format = str(value).strip()
    match_obj = re.match('https?://', value_format, re.IGNORECASE)
    if not match_obj:
        return 2
    try:
        session = requests.Session()
        session.mount('http://', HTTPAdapter(max_retries=3))
        session.mount('https://', HTTPAdapter(max_retries=3))
        response = session.get(url=value, timeout=(3, 6), headers=REQUEST_HEADERS, verify=False, proxies=PROXIES)
        logger.info(f'请求链接：{value_format}，请求状态：{response.status_code}')
        return response.ok
    except Exception as e:
        logger.debug(f'请求链接：{value_format}，请求失败，失败原因：{e}')
        return False


def cell_processor(cell, check_res):
    """
    根据审查结果标记cell
    :param cell: 单元格
    :param check_res: 审查结果
    :return:
    """
    if not check_res:
        # if True:
        try:
            cell.style = 'highlight'
            # cell.style = highlight
        except Exception as e:
            try:
                cell.style = get_highlight_style()
            except Exception as e1:
                logger.debug(f'E1-设置样式对象失败：{e1}')
            logger.debug(f'E-设置样式对象失败：{e}')


def read_excel(file_path) -> Workbook:
    """
    读取excel文件
    :param file_path: inputFile内的相对路径
    :return: sheet
    """
    workbook_ = load_workbook(file_path)
    return workbook_


def get_max_row(sheet) -> int:
    """
    计算最大行数
    :param sheet: 工作表
    :return: 最大行数
    """
    rows_ = sheet.iter_rows()
    # 当前检测的行号
    current_row_index = 0
    # 上次的非空行号
    last_nblank_row_index = 0
    # 最大检测多余空白行数
    max_check_line_count = 10
    for row in rows_:
        current_row_index += 1
        for cell in row:
            if cell.value:
                last_nblank_row_index = current_row_index
                break
        if current_row_index - last_nblank_row_index > max_check_line_count:
            break
    return last_nblank_row_index


def center_thread(cell, checker, processor):
    """
    核心处理逻辑抽取，用于多线程
    :param cell: 单元格
    :param checker: 检测函数
    :param processor: 处理函数
    :return: None
    """
    # cell内容
    cell_value = cell.value

    # # cell编号
    # cell_index = str(cell.column_letter) + str(cell.col_idx)

    # checker进行判断
    check_res = checker(cell_value)
    if check_res == 2:
        return

    # processor进行处理
    processor(cell, check_res)
    # logger.info(f'当前单元格：{cell_index}')
    # time.sleep(0.011)


def check_and_label_cells(sheet, checker, processor):
    """
    检查cell并标记
    :param highlight: 标记对象
    :param processor: 处理工具
    :param checker: 审查工具
    :param sheet: 活动表
    :return:
    """
    # 获取所有row
    rows = sheet.iter_rows()

    max_row_count = get_max_row(sheet)
    logger.info(f'excel文件最大行数为：{max_row_count}')
    # 获取所有cell
    row_index = 0
    for row in rows:
        for cell in row:
            task = tpe.submit(center_thread, cell, checker, processor)
            task_list.append(task)
            # center_thread(cell, checker, processor)
            # # cell内容
            # cell_value = cell.value
            #
            # # # cell编号
            # # cell_index = str(cell.column_letter) + str(cell.col_idx)
            #
            # # checker进行判断
            # check_res = checker(cell_value)
            # if check_res == 2:
            #     continue
            #
            # # processor进行处理
            # processor(cell, check_res, highlight)
            # # logger.info(f'当前单元格：{cell_index}')
            # # time.sleep(0.011)
        if row_index > max_row_count:
            return
        row_index = row_index + 1
        logger.info(f'当前行数：{row_index}')

    wait(task_list, return_when=ALL_COMPLETED, timeout=10)


def run(filename) -> str:
    set_configuration()
    make_filedir()
    input_filepath, output_filepath, log_filepath = get_filepath(filename)
    logger.add(log_filepath, compression="zip")
    highlight_obj = get_highlight_style()
    workbook = read_excel(input_filepath)
    worksheet = workbook.worksheets[0]
    check_and_label_cells(worksheet, cell_checker, cell_processor)
    # tpe.shutdown(wait=True)
    workbook.save(output_filepath)
    logger.info(f'本次生成文件为：{output_filepath}')
    return output_filepath


if __name__ == '__main__':
    start_time = time.time()
    path = sys.argv[1]
    # path = 'BH_2.xlsx'
    # path = 'elf_3.xlsx'
    # path = 'BH.xlsx'
    output_filepath = run(path)
    output_filepath_absolute = os.path.abspath(output_filepath)
    end_time = time.time()
    logger.info(f'标记结束，耗时：{end_time - start_time}')
    os.startfile(output_filepath_absolute)
