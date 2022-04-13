import os

import requests
from openpyxl import load_workbook


# REQUEST_HEADERS = {
#     'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
#                   'Chrome/100.0.4896.60 Safari/537.36 Edg/100.0.1185.29 ',
#     # 'Host': 'zhannei.baidu.com',
#     'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
#     'accept-encoding': 'gzip, deflate, br',
#     # 'content-type': 'image/jpeg'
# }
# url = 'https://www.elfcosmetics.com/dw/image/v2/BBXC_PRD/on/demandware.static/-/Sites-elf-master/default/dwcbcbee5f/2021/81675_BRFAC_OpenA_R.jpg?sw=600&sh=600&sm=fit&sfrm=png'
# url = 'https://click.linksynergy.com/link?id=o0PH*nrGc0U&offerid=570454.10&type=15&murl=https%3A%2F%2Fwww.bhcosmetics.com%2Fproducts%2Fcashmere-cream%3Fvariant%3D31782995263524'
# response = requests.get(url=url, headers=REQUEST_HEADERS)
# text = response.text
# code = response.status_code
# print(text)
# print(code)

def get_max_row(sheet) -> int:
    """
    计算最大行数
    :param sheet: 工作表
    :return: 最大行数
    """
    rows_ = sheet.iter_rows()
    # 当前检测的行号
    current_row_index = 0
    # 上次的非空行号
    last_nblank_row_index = 0
    # 最大检测多余空白行数
    max_check_line_count = 10
    for row in rows_:
        current_row_index += 1
        for cell in row:
            if cell.value:
                last_nblank_row_index = current_row_index
                break
        if current_row_index - last_nblank_row_index > max_check_line_count:
            break
    return last_nblank_row_index


# cell = sheet.cell(row=num, column=1).value
# if cell:
#     num = 1 + num
#     print(num)
# else:
#     return num


path = 'inputFile/elf_1.xlsx'
# workbook_ = load_workbook(path, data_only=True)
# worksheet_ = workbook_.worksheets[0]
# print(worksheet_)
# max_rowcount = get_max_row(worksheet_)
# print(max_rowcount)
# rows = worksheet_.iter_rows()
# cells = worksheet_.active_cell

# max_row = get_max_row(worksheet_)
# min_row = worksheet_.min_row
# print(min_row)
# max_row = worksheet_.max_row
# print(max_row)
# c = worksheet_.columns
#
# while cc := next(c):
#     print(cc)
# rows_ = worksheet_.iter_rows(max_col=27)
# row = next(rows_)
# print(row)
# max_row = get_max_row(worksheet_)
# print(max_row)

i = 'sample'
s = os.mkdir(i)
print(s)
